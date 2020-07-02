import time
import re
import requests
from bs4 import BeautifulSoup as bs
from selenium import webdriver
from selenium.common import exceptions
from selenium.webdriver.chrome.options import Options

# chromedriver 위치
chrome_options = Options()  # 한번만 사용할 것이기 때문에 self를 붙이지 않음
chrome_options.add_argument("--headless")  # CLI (User-agent)
driver = webdriver.Chrome(chrome_options=chrome_options,
                      executable_path="C:\python_source\section3\chromedriver.exe")
wd = 'C:\python_source\section3\chromedriver.exe'
query = '이낙연'
startdate = '20180201'
enddate = '20180228'
urls_list = []
#v페이지수 결정
pagenum = 30
for page in range(1, pagenum+1):

    # 크롤링하고자 하는 사이트
    url = 'https://search.daum.net/search?w=news&q=' + query + '&DA=STC&sd=' + startdate + '000000&ed=' + enddate + '235959&period=u' +'&p=' + str(page)
    web = requests.get(url).content
    source = bs(web, 'html.parser')
   

    for urls in source.find_all('a', {'class': "f_link_b"}):
        new_url_content = urls["href"][-21:-4]
        new_url = "http://news.v.daum.net/v/"+new_url_content

        urls_list.append(new_url)

print(urls_list)

# 크롬드라이버 로드
driver = webdriver.Chrome(wd)

# 매개변수의 주소 띄움
content_list = []

row=1
import openpyxl

wb = openpyxl.Workbook()
sheet = wb['Sheet']

#range = 기사 수
for i in range(len(urls_list)):
    driver.get(urls_list[i])
    print(driver.current_url)

    time.sleep(5)
    try:
        while driver.find_element_by_xpath('//*[@id="alex-area"]/div/div/div/div[3]/div[2]/a').text !='':
            driver.find_element_by_xpath('//*[@id="alex-area"]/div/div/div/div[3]/div[2]/a').click()
            time.sleep(4)

    except:
        pass

    time.sleep(3)

    pages = 0  # 한 페이지당 약 20개의 댓글이 표시
    try:
        while True:  # 댓글 페이지가 몇개인지 모르므로.
            if not driver.find_element_by_css_selector("#alex-area > div > div > div > div.cmt_box > div.alex_more > a").click():
                pass
            else:
                driver.find_element_by_css_selector("#alex-area > div > div > div > div.cmt_box > div.alex_more > a").click()
                time.sleep(4)
                # print(pages, end=" ")
                # pages += 1

    except exceptions.ElementNotVisibleException as e:  # 페이지 끝
        pass

    except Exception as e:  # 다른 예외 발생시 확인
        print(e)

    html = driver.page_source
    dom = bs(html, "lxml")

    # 댓글이 들어있는 페이지 전체 크롤링
    comments_raw = dom.find_all("p", "desc_txt font_size_17")
    # print(comments_raw)
    #     print(i,end='\n\n')



    # 댓글의 text만 뽑는다.
    # comments = [comment.text for comment in comments_raw]
    for comment in comments_raw:
        # sheet.append
        print('\n<=댓글=>')
        comit=comment.text
        print(comit)
        sheet.cell(row=row ,column=1).value = comit
        # sheet.append(comit)

        row = row + 1
        wb.save('C:/python_source/section3/daum_news_crawler/201802다음뉴스댓글.xlsx')
    wb.close()


    # import csv
    # with open('C:\python_source\section3\daum_news_crawler\다음뉴스댓글.csv','w', newline='')as f:
    #     makewrite = csv.writer(f)
    # # ('C:\python_source\section3\daum_news_crawler\다음뉴스댓글.csv', encoding='cp949')

    # while True:
        # sheet.cell(row=row, column=0).value = comments[com]
    # for com in comments:
    #     print(com)

        #     print(i,end='\n\n')


    # if not comments:
    #     break

# #range = 최대 댓글 수
#     for k in range(100):
#         try:
#             contents = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div[2]/div[1]/div[3]/div[7]/div/div/div/div/div[3]/ul[2]/li[' +str(k)+']/div/p').text
#             contents = contents.replacce("\n", " ")
#             print(contents)
#             content_list.append(contents)
#
#             print(content_list)
#         except:
#             pass
#         time.sleep(5)
#
# print(content_list)
#
# #
# pages = 0  # 한 페이지당 약 20개의 댓글이 표시
# try:
#     while True:  # 댓글 페이지가 몇개인지 모르므로.
#         driver.find_element_by_css_selector("#alex-area > div > div > div > div.cmt_box > div.alex_more > a").click()
#         time.sleep(6)
#         print(pages, end=" ")
#         pages += 1
#
# except exceptions.ElementNotVisibleException as e:  # 페이지 끝
#     pass
#
# except Exception as e:  # 다른 예외 발생시 확인
#     print(e)
#
# html = driver.page_source
# dom = bs(html, "lxml")
#
# # 댓글이 들어있는 페이지 전체 크롤링
# comments_raw = dom.find_all("p", "desc_txt font_size_17")
#
# # 댓글의 text만 뽑는다.
# comments = [comment.text for comment in comments_raw]
# for i in comments:
#     print(i)
